import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
import os
import PyPDF2
from docx import Document
from openpyxl import load_workbook

# Function to search meaning online with retries
def search_online(word):
    original_word = word
    if word.endswith('s'):
        word = word[:-1]  # Remove the 's' at the end
    
    url = f"https://www.english-bangla.com/dictionary/{word}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    for attempt in range(3):  # Retry up to 3 times
        try:
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code == 404:
                return f"{original_word} = not found online"
            else:
                content = response.text
                soup = BeautifulSoup(content, 'html.parser')
                span_tags = soup.find_all('span', class_='format1')
                
                if not span_tags:
                    alt_meaning_tag = soup.find('span', class_='meaning')
                    if alt_meaning_tag:
                        alt_meaning_text = alt_meaning_tag.text.strip()
                        return f"{original_word} = {alt_meaning_text.split(' ', 1)[-1].strip()}"
                    else:
                        return search_alternate_online(original_word)
                else:
                    meanings = [span.text.strip().split(' ', 1)[-1].strip() for span in span_tags]
                    return f"{original_word} = {'; '.join(meanings)}"
        except requests.RequestException as e:
            if attempt < 2:
                continue
            else:
                return f"{original_word} = error {e}"

# Function to search meaning online on an alternate site with retries
def search_alternate_online(word):
    url = f"https://www.shabdkosh.com/search-dictionary?lc=bn&sl=en&tl=bn&e={word}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    for attempt in range(3):  # Retry up to 3 times
        try:
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code != 200:
                return f"{word} = not found online"
            
            content = response.text
            soup = BeautifulSoup(content, 'html.parser')
            meaning_tags = soup.find_all('li', class_='d-flex justify-content-between flex-wrap ps-3 mn-item')
            
            if not meaning_tags:
                return f"{word} = not found"
            
            meanings = [tag.find('a', class_='e in l ms-2').text.strip() for tag in meaning_tags]
            return f"{word} = {'; '.join(meanings)}"
        except requests.RequestException as e:
            if attempt < 2:
                continue
            else:
                return f"{word} = error {e}"

# Function to extract text from URLs
def extract_words_from_urls(urls):
    words = set()
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    def get_words_from_url(url):
        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            text = soup.get_text()
            return text.split()
        except requests.RequestException:
            return []

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(get_words_from_url, url): url for url in urls}
        for future in as_completed(futures):
            words.update(future.result())
    
    return list(words)

# Function to extract text from .txt, .xlsx, .docx, .pdf files
def extract_words_from_files(file_paths):
    words = set()
    
    for file_path in file_paths:
        if file_path.endswith('.txt'):
            with open(file_path, 'r', encoding='utf-8') as file:
                text = file.read()
                words.update(text.split())
        
        elif file_path.endswith('.xlsx'):
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        words.update(str(cell).split())
        
        elif file_path.endswith('.docx'):
            doc = Document(file_path)
            for para in doc.paragraphs:
                words.update(para.text.split())
        
        elif file_path.endswith('.pdf'):
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page in range(len(reader.pages)):
                    text = reader.pages[page].extract_text()
                    words.update(text.split())
    
    return list(words)

# Main function to handle word search
def search_meanings(words, urls=None, file_paths=None):
    allowed_chars = set('abcdefghijklmnopqrstuvwxyz ')
    
    if urls:
        url_words = extract_words_from_urls(urls)
        words.extend(url_words)

    if file_paths:
        file_words = extract_words_from_files(file_paths)
        words.extend(file_words)

    # Filter input text based on allowed characters, replacing disallowed characters with spaces
    filtered_words = []
    for word in words:
        filtered_word = ''.join(char if char in allowed_chars else ' ' for char in word.lower())
        filtered_word = ' '.join(filtered_word.split())  # Remove extra spaces
        if filtered_word:  # Only add non-empty words
            filtered_words.extend(filtered_word.split())  # Split into individual words
    unique_words = list(set(filtered_words))  # Remove duplicates

    if not unique_words:
        print("No valid words to search.")
        return

    print(f"Total input unique words: {len(unique_words)}")

    # Load previously saved not found results
    saved_not_found_results = set()
    try:
        with open(r"C:\Users\style\Desktop\10 july py\TEST 2 N F.txt", "r", encoding="utf-8") as file:
            for line in file:
                saved_not_found_results.add(line.strip())
    except FileNotFoundError:
        pass

    # Load previously saved found results
    saved_found_results = set()
    try:
        with open(r"C:\Users\style\Desktop\10 july py\TEST 1 F W.txt", "r", encoding="utf-8") as file:
            for line in file:
                saved_found_results.add(line.strip().split('=')[0].strip())
    except FileNotFoundError:
        pass

    # Filter out words already in saved results
    words_to_search = [word for word in unique_words if word not in saved_not_found_results and word not in saved_found_results]

    if not words_to_search:
        print("No new words to search.")
        return

    results = set()
    not_found_results = set()
    error_count = 0

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(search_online, word): word for word in words_to_search}
        for future in tqdm(as_completed(futures), total=len(futures), desc="Searching", unit="word"):
            result = future.result()
            if "not found" in result:
                not_found_results.add(result.split('=')[0].strip())
            elif "error" in result:
                error_count += 1
                with open("C:\\Users\\style\\Desktop\\10 july py\\hieror.txt", "a", encoding="utf-8") as error_file:
                    error_file.write(result + "\n")
            else:
                results.add(result)

    # Print search results
    for result in results:
        print(result)

    # Update saved not found results with new unique words
    with open(r"C:\Users\style\Desktop\10 july py\TEST 2 N F.txt", "a", encoding="utf-8") as file:
        for result in not_found_results:
            if result not in saved_not_found_results:
                file.write(result + "\n")

    # Update saved found results with new unique words
    with open(r"C:\Users\style\Desktop\10 july py\TEST 1 F W.txt", "a", encoding="utf-8") as file:
        for result in results:
            word = result.split('=')[0].strip()
            if word not in saved_found_results:
                file.write(result + "\n")

    # Print summary
    print(f"Total input words: {len(unique_words)}")
    print(f"Words found online: {len(results)}")
    print(f"Words not found online: {len(not_found_results)}")
    print(f"Errors encountered : {error_count}")

    # Input text directly via code
input_text = """
girl = ; কুমারী মেয়ে
baby = শিশু
phoneds = করা;
phone = দূরভাস
said = ব্যক্ত;
fan  fanaaa

"""
words = input_text.split()

# List of URLs to fetch words from
urls = [
# "https://www.thedailystar.net/news/bangladesh/news/quota-reform-demo-protesters-set-resume-blockade-today-3653081",
# # Add more URLs as needed
]

# List of file paths to fetch words from such as, .txt, xlsx, .pdf, .docx
file_paths = [

# r"C:\Users\style\Desktop\10 july py\check.txt"
# r"C:\Users\style\Documents\check.docx"
# r"H:\locks\youtube thumbnails dorker\Google Classroom Code_ Fall 2022_TE. PC.xlsx"
# r"H:\locks\German Language dorker\German Documents\Way To Germany  -  A Complete Guide For Masters Students.pdf"

]

# Start the word search
search_meanings(words, urls, file_paths)
